"""
Excel export module for invoice data.

Handles:
- Writing invoice data to Excel files
- ZAR currency formatting
- Multi-row line item export
- Template creation
"""

import logging
from datetime import datetime
from pathlib import Path
from typing import Optional, Union

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.invoice import ExtractedInvoice

logger = logging.getLogger(__name__)


class ExcelExporter:
    """
    Exports invoice data to Excel files.
    
    Features:
    - ZAR currency formatting (R #,##0.00)
    - Each line item as a separate row
    - Shared header data on each row
    - Professional formatting
    - Append to existing file or create new
    """
    
    # Column headers for the export
    HEADERS = [
        "Invoice Number",
        "Invoice Date",
        "Due Date",
        "Seller Name",
        "Seller VAT Number",
        "Seller Address",
        "Customer Name",
        "Customer Account",
        "Customer Address",
        "Item Description",
        "Quantity",
        "Unit",
        "Unit Price (R)",
        "Line Total (R)",
        "VAT Rate (%)",
        "Subtotal (R)",
        "VAT Amount (R)",
        "Total Due (R)",
        "Extraction Date",
        "Confidence Score",
    ]
    
    # Column widths
    COLUMN_WIDTHS = {
        "A": 15,  # Invoice Number
        "B": 12,  # Invoice Date
        "C": 12,  # Due Date
        "D": 25,  # Seller Name
        "E": 15,  # Seller VAT
        "F": 30,  # Seller Address
        "G": 25,  # Customer Name
        "H": 15,  # Customer Account
        "I": 30,  # Customer Address
        "J": 35,  # Item Description
        "K": 10,  # Quantity
        "L": 10,  # Unit
        "M": 12,  # Unit Price
        "N": 12,  # Line Total
        "O": 10,  # VAT Rate
        "P": 12,  # Subtotal
        "Q": 12,  # VAT Amount
        "R": 12,  # Total Due
        "S": 12,  # Extraction Date
        "T": 10,  # Confidence
    }
    
    def __init__(self):
        """Initialize the Excel exporter."""
        self._setup_styles()
    
    def _setup_styles(self):
        """Set up Excel styles for formatting."""
        # Header style
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Currency style
        self.currency_format = 'R #,##0.00'
        
        # Percentage style
        self.percentage_format = '0.0%'
        
        # Border style
        thin_border = Side(style="thin", color="CCCCCC")
        self.cell_border = Border(
            left=thin_border,
            right=thin_border,
            top=thin_border,
            bottom=thin_border,
        )
        
        # Alternating row colors
        self.even_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    
    def export(
        self,
        invoice: ExtractedInvoice,
        file_path: Union[str, Path],
        append: bool = True,
    ) -> Path:
        """
        Export invoice data to Excel file.
        
        Args:
            invoice: Extracted invoice data
            file_path: Path to Excel file
            append: If True, append to existing file; if False, create new
            
        Returns:
            Path to the exported file
        """
        file_path = Path(file_path)
        
        # Ensure .xlsx extension
        if file_path.suffix.lower() != ".xlsx":
            file_path = file_path.with_suffix(".xlsx")
        
        # Load existing or create new workbook
        if append and file_path.exists():
            wb = load_workbook(file_path)
            ws = wb.active
            start_row = ws.max_row + 1
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Invoice Data"
            self._write_headers(ws)
            start_row = 2
        
        # Convert invoice to rows
        rows = invoice.to_excel_rows()
        
        # Write data rows
        for i, row_data in enumerate(rows):
            row_num = start_row + i
            self._write_row(ws, row_num, row_data)
            
            # Apply alternating row colors
            if row_num % 2 == 0:
                for col in range(1, len(self.HEADERS) + 1):
                    ws.cell(row=row_num, column=col).fill = self.even_row_fill
        
        # Auto-fit columns (approximate)
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        
        # Save workbook
        wb.save(file_path)
        logger.info(f"Exported {len(rows)} rows to {file_path}")
        
        return file_path
    
    def _write_headers(self, ws):
        """Write header row with formatting."""
        for col, header in enumerate(self.HEADERS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border
        
        # Freeze header row
        ws.freeze_panes = "A2"
    
    def _write_row(self, ws, row_num: int, row_data: dict):
        """Write a single data row."""
        # Map data to columns
        column_mapping = [
            ("invoice_number", None),
            ("invoice_date", None),
            ("due_date", None),
            ("seller_name", None),
            ("seller_vat_number", None),
            ("seller_address", None),
            ("customer_name", None),
            ("customer_account", None),
            ("customer_address", None),
            ("item_description", None),
            ("quantity", None),
            ("unit_of_measure", None),
            ("unit_price", self.currency_format),
            ("line_total", self.currency_format),
            ("vat_rate", None),  # Will format as percentage
            ("subtotal", self.currency_format),
            ("vat_amount", self.currency_format),
            ("total_due", self.currency_format),
            ("extraction_date", None),
            ("confidence_score", None),
        ]
        
        for col, (key, number_format) in enumerate(column_mapping, start=1):
            value = row_data.get(key)
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = self.cell_border
            
            # Apply number format for currency columns
            if number_format and value is not None:
                cell.number_format = number_format
            
            # Format VAT rate as percentage
            if key == "vat_rate" and value is not None:
                cell.value = value / 100  # Convert to decimal for percentage format
                cell.number_format = "0%"
    
    def create_template(self, file_path: Union[str, Path]) -> Path:
        """
        Create an empty Excel template with headers.
        
        Args:
            file_path: Path for the template file
            
        Returns:
            Path to the created template
        """
        file_path = Path(file_path)
        
        # Ensure .xlsx extension
        if file_path.suffix.lower() != ".xlsx":
            file_path = file_path.with_suffix(".xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Data"
        
        # Write headers
        self._write_headers(ws)
        
        # Set column widths
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        
        # Add data validation notes
        ws["A2"] = "Enter invoice data here..."
        ws["A2"].font = Font(italic=True, color="808080")
        
        wb.save(file_path)
        logger.info(f"Created template at {file_path}")
        
        return file_path
    
    def export_multiple(
        self,
        invoices: list[ExtractedInvoice],
        file_path: Union[str, Path],
    ) -> Path:
        """
        Export multiple invoices to a single Excel file.
        
        Args:
            invoices: List of extracted invoices
            file_path: Path to Excel file
            
        Returns:
            Path to the exported file
        """
        file_path = Path(file_path)
        
        if file_path.suffix.lower() != ".xlsx":
            file_path = file_path.with_suffix(".xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Invoice Data"
        
        # Write headers
        self._write_headers(ws)
        
        # Write all invoices
        current_row = 2
        for invoice in invoices:
            rows = invoice.to_excel_rows()
            for row_data in rows:
                self._write_row(ws, current_row, row_data)
                if current_row % 2 == 0:
                    for col in range(1, len(self.HEADERS) + 1):
                        ws.cell(row=current_row, column=col).fill = self.even_row_fill
                current_row += 1
        
        # Set column widths
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        
        wb.save(file_path)
        total_rows = current_row - 2
        logger.info(f"Exported {len(invoices)} invoices ({total_rows} rows) to {file_path}")
        
        return file_path


def export_invoice(
    invoice: ExtractedInvoice,
    file_path: Union[str, Path],
    append: bool = True,
) -> Path:
    """
    Convenience function to export an invoice.
    
    Args:
        invoice: Extracted invoice data
        file_path: Path to Excel file
        append: Whether to append to existing file
        
    Returns:
        Path to exported file
    """
    exporter = ExcelExporter()
    return exporter.export(invoice, file_path, append)


def create_excel_template(file_path: Union[str, Path]) -> Path:
    """
    Convenience function to create an Excel template.
    
    Args:
        file_path: Path for the template
        
    Returns:
        Path to created template
    """
    exporter = ExcelExporter()
    return exporter.create_template(file_path)
